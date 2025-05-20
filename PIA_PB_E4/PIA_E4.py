from openpyxl import Workbook, load_workbook
from collections import Counter
import re
import matplotlib.pyplot as plt

#Reorganización de datos
artista = {"nombre" : "",
           "id" : "",
           "genero": [],
           "seguidores" : 0,
           "popularidad": 0}

evento = {"dia 1" : "NA",
          "dia 2" : "NA",
          "dia 3" : "NA"}

lista_dia1 = []
lista_dia2 = []
lista_dia3 = []
lista_generos = []
lista_seguidores = []
lista_popularidad = []
lista_dia1_generos = []
lista_dia2_generos = []
lista_dia3_generos = []
lista_dia1_seguidores = []
lista_dia2_seguidores = []
lista_dia3_seguidores = []
lista_dia1_popularidad = []
lista_dia2_popularidad = []
lista_dia3_popularidad = []

i=1
a=1
b=0

#Lectura de datos
with open ("PIA_E2_ARCHIVO.txt","r") as file:
    for line in file:
        info = line.split("}]")
        evento["dia 1"] = info[0]
        evento["dia 2"] = info[1]
        evento["dia 3"] = info[2]
        #Validación de datos
        for dia in range(3):
            while True:
                infor = evento[f"dia {a}"].split(":")
                patron_nombre = re.compile(r"\"?\\?\'?([a-zA-Z0-9É\s]+)\"?\'?\\?\, (\'id\')")
                mo1 = patron_nombre.search(infor[i])
                artista["nombre"] = mo1.group(1)
                patron_id = re.compile(r"\'(\w+)\'\, (\'generos\')")
                mo2 = patron_id.search(infor[i+1])
                artista["id"] = mo2.group(1)
                patron_generos = re.compile(r"(\[?\D+\]?)\, (\'seguidores\')")
                mo3 = patron_generos.search(infor[i+2])
                artista["genero"] = mo3.group(1)
                patron_seguidores = re.compile(r"(\d+)\, (\'popularidad\')")
                mo4 = patron_seguidores.search(infor[i+3])
                artista["seguidores"] = mo4.group(1)
                patron_popularidad = re.compile(r"(\d+)\}\]?\]?\,? \{?(\'nombre\')?")
                mo5 = patron_popularidad.search(infor[i+4])
                if mo5 == None:
                    artista["popularidad"] = 93
                else:
                    artista["popularidad"] = mo5.group(1)
                datos = artista.copy()
                if dia==0:
                    lista_dia1.append(datos)
                elif dia==1:
                    lista_dia2.append(datos)
                else:
                    lista_dia3.append(datos)
                i += 5
                if mo1.group(1) == 'Travis Scott':
                    a += 1
                    i=1
                    break

#Preparación de visualización
lista_evento = [lista_dia1, lista_dia2, lista_dia3]                
for dia in evento:
    for artista in lista_evento[b]:
        gen = artista["genero"].split(",")
        for elemento in gen:
            patron_gen = re.compile(r"\[?(\"|')([a-zA-Zá-úñ\s\-\&]+)(\"|')\]?")
            mo6 = patron_gen.search(elemento)
            if b==0:
                lista_dia1_generos.append(mo6.group(2))
            elif b==1:
                lista_dia2_generos.append(mo6.group(2))
            else:
                lista_dia3_generos.append(mo6.group(2))
        if b==0:
            lista_dia1_seguidores.append(artista["seguidores"])
            lista_dia1_popularidad.append(artista["popularidad"])
        elif b==1:
            lista_dia2_seguidores.append(artista["seguidores"])
            lista_dia2_popularidad.append(artista["popularidad"])
        else:
            lista_dia3_seguidores.append(artista["seguidores"])
            lista_dia3_popularidad.append(artista["popularidad"])
    b +=1

lista_generos = [lista_dia1_generos, lista_dia2_generos, lista_dia3_generos]
lista_seguidores = [lista_dia1_seguidores, lista_dia2_seguidores,lista_dia3_seguidores]
lista_popularidad = lista_dia1_popularidad+lista_dia2_popularidad+lista_dia3_popularidad

evento = []
for lista in lista_evento:
    for artista in lista:
        evento.append(artista)



#calculos estadisticos
#SUMA SEGUDORES
lista_dia1_seguidores = [int(c) for c in lista_dia1_seguidores] #esto viene en la documentacion de python
lista_dia2_seguidores = [int(c) for c in lista_dia2_seguidores]
lista_dia3_seguidores = [int(c) for c in lista_dia3_seguidores]


suma_dia1 = sum(lista_dia1_seguidores)
suma_dia2 = sum(lista_dia2_seguidores)
suma_dia3 = sum(lista_dia3_seguidores)
suma_coachella = suma_dia1 + suma_dia2 + suma_dia3
print("la suma de seguidores del dia 1 es: ", suma_dia1)
print("la suma de seguidores del dia 2 es: ", suma_dia2)
print("la suma de seguidores del dia 3 es: ", suma_dia3)
print("la suma total de seguidores:", suma_coachella) #con este dato podemos estimar la cantidad total de asistentes al festival.

#Frecuencia DE los GENEROS
frec_dia1 = Counter(lista_dia1_generos)
frec_dia2 = Counter(lista_dia2_generos)
frec_dia3 = Counter(lista_dia3_generos)

generos = []
for lista in lista_generos:
    for genero in lista:
        generos.append(genero)
        
frec_coachella = Counter(generos)
print("los 5 generos mas escuchados en el dia 1:")
print('{:<20}\t{:>12}'.format("Genero","Apariciones"))
for genero in frec_dia1.most_common(5):
    print('{:<20}\t{:>12}'.format(genero[0],genero[1]))

print("------------")

print("los 5 generos mas escuchados en el dia 2: ")
print('{:<20}\t{:>12}'.format("Genero","Apariciones"))
for genero in frec_dia2.most_common(5):
    print('{:<20}\t{:>12}'.format(genero[0],genero[1]))

print("------------")

print("los 5 generos mas escuchados en el dia 3: ")
print('{:<20}\t{:>12}'.format("Genero","Apariciones"))
for genero in frec_dia3.most_common(5):
    print('{:<20}\t{:>12}'.format(genero[0],genero[1]))


print("------------")

print("los 5 generos mas escuchados en COACHELLA: ")
print('{:<20}\t{:>12}'.format("Genero","Apariciones"))
for genero in frec_coachella.most_common(5):
    print('{:<20}\t{:>12}'.format(genero[0],genero[1]))

#DATOS GUARDADOS EN ARCHIVO

fo = open("CALCULOS_PIA_E3.txt", "w")
fo.write("Los 5 géneros más escuchados en el dia 1")
fo.write(str(frec_dia1.most_common(5)))

fo.write("Los 5 géneros más escuchados en el día 2:")
fo.write(str(frec_dia2.most_common(5)))

fo.write("Los 5 géneros más escuchados en el día 3:")
fo.write(str(frec_dia3.most_common(5)))

fo.write("Los 5 géneros más escuchados en COACHELLA:")
fo.write(str(frec_coachella.most_common(5)))



fo.write(f"\nSuma total de seguidores día 1: {suma_dia1}\n")
fo.write(f"Suma total de seguidores día 2: {suma_dia2}\n")
fo.write(f"Suma total de seguidores día 3: {suma_dia3}\n")
fo.write(f"Suma total de seguidores del festival: {suma_coachella}\n")

fo.close()

#VALIDACIÓN DE DATOS PARA VISUALIZACIÓN
seguidores_lineal_y=[suma_dia1,suma_dia2,suma_dia3]
seguidores_lineal_x=["Dia1","Dia2","Dia3"]
seguidores_pie_y=[suma_dia1,suma_dia2,suma_dia3]
seguidores_pie_label=["Dia1","Dia2","Dia3"]
hist=[int(c) for c in lista_popularidad]
popularidad_hist = hist
generos_barh_y=[]
generos_barh_x=[]
for genero in frec_coachella.most_common(5):
    if genero[0]=="desconocido":
        pass
    else:
        generos_barh_x.append(genero[0])
        generos_barh_y.append(genero[1])

# ALMACENAMIENTO DE DATOS EN EXCEL
try:
    libro = load_workbook("Coachella.xlsx")
except FileNotFoundError:
    libro = Workbook()
    pagina = libro.active
    pagina.title = "Artistas_del festival"

def Artistas_coachella(evento):
    hoja = libro["Artistas_del festival"]

    print("La página activa es:", hoja.title)

    hoja["A1"] = "nombre"
    hoja["B1"] = "id"
    hoja["C1"] = "genero"
    hoja["D1"] = "seguidores"
    hoja["E1"] = "popularidad"

    print("Los encabezados de las celdas son:")
    print(hoja["A1"].value, end="\t")
    print(hoja["B1"].value, end="\t")
    print(hoja["C1"].value, end="\t")
    print(hoja["D1"].value, end="\t")
    print(hoja["E1"].value)

    count = 2
    for artista in evento:
        hoja.cell(count, 1, artista['nombre'])
        hoja.cell(count, 2, artista['id'])
        hoja.cell(count, 3, artista['genero'])
        hoja.cell(count, 4, artista['seguidores'])
        hoja.cell(count, 5, artista['popularidad'])
        count += 1

def seguidores(lista_dia1_seguidores, lista_dia2_seguidores, lista_dia3_seguidores):
    
    hoja = libro.create_sheet("seguidores_coachella")

    print("La página activa es:", hoja.title)

    hoja["A1"] = "seguidores_dia1"
    hoja["B1"] = "seguidores_dia2"
    hoja["C1"] = "seguidores_dia3"

    print("Los encabezados de las celdas son:")
    print(hoja["A1"].value, end="\t")
    print(hoja["B1"].value, end="\t")
    print(hoja["C1"].value)
    count = 2
    for i in range(len(lista_dia1_seguidores)):
        hoja.cell(count, 1, lista_dia1_seguidores[i])
        count += 1

    count = 2
    for i in range(len(lista_dia2_seguidores)):
        hoja.cell(count, 2, lista_dia2_seguidores[i])
        count += 1

    count = 2

    for i in range(len(lista_dia3_seguidores)):
        hoja.cell(count, 3, lista_dia3_seguidores[i])
        count += 1

    

def popularidad(lista_dia1_popularidad, lista_dia2_popularidad, lista_dia3_popularidad):

    hoja = libro.create_sheet("popularidad_coachella")

    print("La página activa es:", hoja.title)

    hoja["A1"] = "Popularidad_dia1"
    hoja["B1"] = "Popularidad_dia2"
    hoja["C1"] = "Popularidad_dia3"

    print("Los encabezados de las celdas son:")
    print(hoja["A1"].value, end="\t")
    print(hoja["B1"].value, end="\t")
    print(hoja["C1"].value)

    count = 2
    for i in range(len(lista_dia1_popularidad)):
        hoja.cell(count, 1, lista_dia1_popularidad[i])
        count += 1

    count = 2
    for i in range(len(lista_dia2_popularidad)):
        hoja.cell(count, 2, lista_dia2_popularidad[i])
        count += 1

    count = 2

    for i in range(len(lista_dia3_popularidad)):
        hoja.cell(count, 3, lista_dia3_popularidad[i])
        count += 1




Artistas_coachella(evento)
seguidores(lista_dia1_seguidores, lista_dia2_seguidores, lista_dia3_seguidores)
popularidad(lista_dia1_popularidad, lista_dia2_popularidad, lista_dia3_popularidad)


print(libro.sheetnames)
libro.save("Coachella.xlsx")

#DISEÑO DE GRÁFICAS
while True:
    menu ="""¿Qué gráfica quieres ver?
1. Gráfica linear de seguidores.
2. Gráfica de pastel para seguidores.
3. Histograma de popularidad.
4. Gráfica de barras horizontales para géneros
5. Salir"""
    print(menu)
    grafica=int(input("ingrese la grafia que desee ver: "))
    if grafica==1:
        plt.plot(seguidores_lineal_x,seguidores_lineal_y, marker='o')
        plt.title("Seguidores de los artísticas presentados en Coachella")
        plt.xlabel("Día del evento")
        plt.ylabel("Cantidad de seguidores en cientos de millones")
        plt.axis(ymin=0,ymax=200000000)
        plt.grid()
        plt.show()
    elif grafica==2:
        plt.pie(seguidores_pie_y,labels=seguidores_pie_label)
        plt.title("Seguidores de los artísticas presentados en Coachella")
        plt.legend(seguidores_pie_y,title="Cantidad de seguidores")
        plt.show()
    elif grafica==3:
        plt.hist(popularidad_hist)
        plt.title("Popularidad de los artistas presentados en Coachella")
        plt.xlabel("Nivel de popularidad de 0-100")
        plt.ylabel("Cantidad de artistas")
        plt.grid()
        plt.show()
    elif grafica==4:
        plt.barh(generos_barh_x,generos_barh_y)
        plt.title("Cantidad de artistas por los 4 géneros musicales más comúnes en Coachella")
        plt.xlabel("Cantidad de artistas")
        plt.ylabel("Género musical")
        plt.axis(xmin=0,xmax=14)
        plt.grid(axis="x")
        plt.show()
    elif grafica==5:
        break
    else:
        grafica=int(input("Este valor no es aceptable, ingresa otro: "))

